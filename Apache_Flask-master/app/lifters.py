import os
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Alignment, NamedStyle
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import datetime
import json
import base64
import requests as RQS
#from boxsdk import OAuth2
#from boxsdk import CCGAuth


def process_excel(json_in):
    file_name = json_in['request_id']
    file_data = json_in['payload']
    message_to_relay = ''

    try:
        destination_folder = os.path.abspath('../temp')
        full_name = destination_folder + '/' + file_name + '.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = 'Flask Python'
        ws.sheet_properties.tabColor = '1072BA'
        ws['A1'] = file_data

        wb.save(full_name)
        wb.close()
        message_to_relay = 'Good'
    except Exception as e:
        message_to_relay = str(e)
    finally:
        if os.path.exists(file_name):
            os.remove(file_name)
        return message_to_relay


# DO NOT USE ANTIQUATED!!! Takes in query first, then convert resultant JSON into Excel file
def process_excel_color(table_id, fid_array_raw, query, new_file_name='color', dest_table_id='', rec_id=1, dest_fid=''):
    try:
        json_in = json.loads(query_qb(table_id, fid_array_raw, query))  # Calling other lifter to query QB
        file_name = new_file_name
        file_columns = json_in['fields']
        file_data = json_in['data']
        message_to_relay = ''

        destination_folder = os.path.abspath('../temp')  # '../temp' when back in lifters, 'temp' in friv
        full_name = destination_folder + '/' + file_name + '.xlsx'

        wb = Workbook()  # Row & Column index start at one
        ws = wb.active
        ws.title = 'Colorized Worksheet'
        ws.sheet_properties.tabColor = '1072BA'

        # Make header row
        for i in range(len(file_columns)):
            column_header = file_columns[i]['label']
            ws.cell(row=1, column=i + 1, value=column_header)

        # Make data rows
        for i in range(len(file_data)):
            for j in range(len(file_columns)):
                dict_in_row = file_data[i]
                dict_val = str(file_columns[j]['id'])
                data_type = str(file_columns[j]['type'])
                cell_value = dict_in_row[dict_val]['value']

                if data_type == "rich-text" and cell_value != "":  # correcting rich-text data
                    font_contents = ""
                    font_color = ""
                    color_code = "00000000"  # default font color to black
                    font_weight = ""
                    ###
                    left_bound = cell_value.find('>')
                    right_bound = cell_value.find('<', 2)
                    ###
                    left_color_bound = cell_value.find('color:')
                    right_color_bound = cell_value.find(';', left_color_bound)
                    if left_color_bound != -1:
                        left_color_bound = left_color_bound + 6
                        font_color = cell_value[left_color_bound + 1:right_color_bound]
                    ###
                    left_weight_bound = cell_value.find('font-weight:')
                    right_weight_bound = cell_value.find(';', left_weight_bound)
                    if left_weight_bound != -1:
                        left_weight_bound = left_weight_bound + 12
                        font_weight = cell_value[left_weight_bound + 1:right_weight_bound]
                    ###
                    cell_value = cell_value[left_bound + 1:right_bound]
                    if font_color != "" or font_weight != "":
                        if font_color != "":
                            if font_color == "blue":
                                color_code = "000000FF"
                            elif font_color == "red":
                                color_code = "00FF0000"
                            else:
                                color_code = "00000000"

                            if font_contents == "":
                                font_contents = "color=" + color_code
                            else:
                                font_contents = font_contents + ",color=" + color_code

                        if font_color != "" and font_weight != "":
                            ws.cell(row=i + 2, column=j + 1, value=cell_value).font = Font(color=color_code, bold=True)
                        elif font_color != "" and font_weight == "":
                            ws.cell(row=i + 2, column=j + 1, value=cell_value).font = Font(color=color_code)
                        else:
                            ws.cell(row=i + 2, column=j + 1, value=cell_value).font = Font(bold=True)

                    else:
                        ws.cell(row=i + 2, column=j + 1, value=cell_value)
                elif data_type == "date" and cell_value != "":
                    temp_list = cell_value.split('-')
                    cell_value = temp_list[1] + "-" + temp_list[2] + "-" + temp_list[0]  # formatting date to MM-DD-YYYY
                    ws.cell(row=i + 2, column=j + 1, value=cell_value)
                else:  # Data type is NOT rich-text
                    ws.cell(row=i + 2, column=j + 1, value=cell_value)

        wb.save(full_name)
        wb.close()

        # Send to destination if applicable
        if dest_table_id != '' and dest_fid != '':
            send_to_skunkworks_x(dest_table_id, rec_id, full_name, file_name + ".xlsx", dest_fid)
        message_to_relay = 'Good'

    except Exception as e:
        message_to_relay = str(e)
        print("Exception!!!: " + message_to_relay)
    finally:
        if os.path.exists(full_name):
            os.remove(full_name)  # Deleting local copy of file
        return message_to_relay


# Runs report for QB via API, then convert resultant JSON into Excel file
def process_excel_color_x(report_id, table_id, qb_realm, qb_token, new_file_name='color', dest_table_id='',
                          dest_rec_id=1, dest_fid=''):
    try:
        json_in = json.loads(run_a_report(report_id, table_id, qb_realm, qb_token))
        last_column = 0
        grouping_present = False
        file_name = new_file_name
        file_columns = json_in['fields']
        file_data = json_in['data']
        message_to_relay = ''

        destination_folder = os.path.abspath('../temp')  # '../temp' when back in lifters, 'temp' in friv
        full_name = destination_folder + '/' + file_name + '.xlsx'

        wb = Workbook()  # Row & Column index start at one
        ws = wb.active
        ws.title = 'Colorized Worksheet'
        ws.sheet_properties.tabColor = '1072BA'

        # Make header row
        for i in range(len(file_columns)):
            column_header = file_columns[i]['label']
            last_column = last_column + 1
            if column_header.find('Grouping ') != -1:
                grouping_present = True
            if column_header.count('_') == 1 and column_header[-1] == '_':
                column_header = column_header[0:-1]
            ws.cell(row=1, column=i + 1, value=column_header)

        # Make data rows
        for i in range(len(file_data)):
            for j in range(len(file_columns)):
                dict_in_row = file_data[i]
                dict_val = str(file_columns[j]['id'])
                data_type = str(file_columns[j]['type'])
                cell_value = dict_in_row[dict_val]['value']

                if data_type == "rich-text" and cell_value != "" and ">N/A<" not in cell_value:  # correcting rich-text date
                    font_color = ""
                    color_code = "00000000"  # default font color to black
                    font_weight = ""
                    ###
                    left_color_bound = cell_value.find('color:')
                    right_color_bound = cell_value.find(';', left_color_bound)
                    if left_color_bound != -1:
                        left_color_bound = left_color_bound + 6
                        font_color = cell_value[left_color_bound + 1:right_color_bound]
                    ###
                    left_weight_bound = cell_value.find('font-weight:')
                    right_weight_bound = cell_value.find(';', left_weight_bound)
                    if left_weight_bound != -1:
                        left_weight_bound = left_weight_bound + 12
                        font_weight = cell_value[left_weight_bound + 1:right_weight_bound]
                    ###
                    if cell_value.count('>') != 0:
                        cell_value = get_cell_value(cell_value)
                    # Removing parentheses (if applicable)
                    if cell_value.find('(') != -1 and cell_value.find(')') != -1 and (len(cell_value) == 12 or len(cell_value) == 10):
                        cell_value = cell_value[1:-1]

                    if font_color != "" or font_weight != "":
                        if font_color != "":
                            if font_color == "blue":
                                color_code = "000000FF"
                            elif font_color == "red":
                                color_code = "00FF0000"
                            else:
                                color_code = "00000000"

                        correct_date = strip_date_from_string(cell_value)
                        cell = ws.cell(row=i + 2, column=j + 1)
                        cell.value = correct_date
                        if len(cell_value) == 8:
                            cell.number_format = 'MM-DD-YY'
                        else:
                            cell.number_format = 'MM-DD-YYYY'

                        if font_color != "" and font_weight != "":
                            cell.font = Font(color=color_code, bold=True)
                        elif font_color != "" and font_weight == "":
                            cell.font = Font(color=color_code)
                        else:
                            cell.font = Font(bold=True)

                    else:
                        correct_date = strip_date_from_string(cell_value)
                        cell = ws.cell(row=i + 2, column=j + 1)
                        cell.value = correct_date
                        if len(cell_value) == 8:
                            cell.number_format = 'MM-DD-YY'
                        else:
                            cell.number_format = 'MM-DD-YYYY'
                elif data_type == "date" and cell_value != "":
                    correct_date = datetime.datetime.strptime(cell_value, "%Y-%m-%d")
                    cell = ws.cell(row=i + 2, column=j + 1)
                    cell.value = correct_date
                    cell.number_format = 'MM-DD-YYYY'
                elif data_type == "rich-text" and ">N/A<" in cell_value:
                    ws.cell(row=i + 2, column=j + 1, value="N/A")
                else:  # Data type doesn't require processing
                    ws.cell(row=i + 2, column=j + 1, value=cell_value)

        # Delete last column if it has grouping
        if grouping_present:
            ws.delete_cols(last_column, 1)

        wb.save(full_name)
        wb.close()

        # Send to destination if applicable
        if dest_table_id != '' and dest_fid != '':
            if qb_token:
                send_to_skunkworks_x(dest_table_id, dest_rec_id, full_name, file_name + ".xlsx", dest_fid, qb_token)
            else:
                send_to_skunkworks_x(dest_table_id, dest_rec_id, full_name, file_name + ".xlsx", dest_fid)
        message_to_relay = 'Good'

    except Exception as e:
        message_to_relay = str(e)
        print("Exception!!!: " + message_to_relay)
    finally:
        if os.path.exists(full_name):
            os.remove(full_name)  # Deleting local copy of file
        return message_to_relay


def process_dummy_template(json_in):
    record_id = json_in['request_id']
    file_name = json_in['name']
    file_data = json_in['payload']
    file_data_b = json_in['payload_b']

    try:
        template_path = os.path.abspath('../doc_templates')
        wb = load_workbook(template_path + '/invoice_template.xlsx')
        destination_folder = os.path.abspath('../temp')
        full_name = destination_folder + '/' + file_name + '.xlsx'

        ws = wb.active
        ws.title = 'Flask Python Generated'
        ws.sheet_properties.tabColor = '1072BA'

        ws['A4'] = 'Name:'
        ws['B4'] = file_data

        ws['A10'] = 'Value:'
        ws['B10'] = file_data_b

        wb.save(full_name)
        wb.close()
        return 'Good'
    except Exception as e:
        return str(e)


def web_scrape(target_url):
    """Pulling data from other sites"""
    return


def query_qb(table_id, fid_array_raw, query):
    message_to_relay = 'Good'

    try:
        # was a single number or chain of numbers input?
        if fid_array_raw.isnumeric():
            fid_array = [int(fid_array_raw)]
        else:
            fid_string_array = fid_array_raw.split(',')
            fid_array = [int(numeric_string) for numeric_string in fid_string_array]  # int array from string array

        headers = {
            'QB-Realm-Hostname': 'SECRET SECRET SECRET',
            'User-Agent': '{User-Agent}',
            'Authorization': 'QB-USER-TOKEN SECRET SECRET SECRET'
        }
        body = {"from": table_id, "select": fid_array, "where": query}  # "where": "{FID.EX.'2'}"
        r = RQS.post(
            'https://api.quickbase.com/v1/records/query',
            headers=headers,
            json=body
        )
        # PING LOCALHOST: http://127.0.0.1:5000/ping/qb_query?table_id=bnvqs469i&fid_array=8,9&criteria={%2763%27.XEX.%272%27}
        message_to_relay = json.dumps(r.json(), indent=4)
    except Exception as e:
        message_to_relay = str(e)
    finally:
        return message_to_relay


def run_a_report(report_id, table_id, qb_realm='SECRET SECRET SECRET',
                 auth_token='SECRET SECRET SECRET'):
    message_to_relay = 'Good'
    final_auth_token = 'QB-USER-TOKEN ' + auth_token
    target_url = 'https://api.quickbase.com/v1/reports/' + str(report_id) + '/run'

    try:
        headers = {
            'QB-Realm-Hostname': qb_realm,
            'User-Agent': '{User-Agent}',
            'Authorization': final_auth_token
        }
        params = {'tableId': table_id}
        r = RQS.post(
            target_url,
            headers=headers,
            params=params
        )
        # PING LOCALHOST: http://127.0.0.1:5000/ping/qb_query?table_id=bnvqs469i&fid_array=8,9&criteria={%2763%27.XEX.%272%27}
        message_to_relay = json.dumps(r.json(), indent=4)
    except Exception as e:
        message_to_relay = str(e)
    finally:
        return message_to_relay


def send_to_skunkworks(tab_id, rec_id, file_name):
    table_id = str(tab_id)
    record_id = int(rec_id)
    message_to_relay = ''

    try:
        file_byte_content = open(file_name, 'rb').read()  # read file as byte stream
        base_64_bytes = base64.b64encode(file_byte_content)  # convert byte into "base64 byte"
        base64_string = base_64_bytes.decode('utf-8')  # convert base54 byte stream into utf-8 encoded string

        headers = {
            'QB-Realm-Hostname': 'SECRET SECRET SECRET',
            'User-Agent': '{User-Agent}',
            'Authorization': 'SECRET SECRET SECRET'
        }

        body = {
            'to': table_id,
            'data': [
                {
                    '3': {
                        'value': record_id
                    },
                    '7': {
                        'value': {
                            'fileName': 'excel.xlsx',
                            'data': base64_string
                        }
                    }
                }
            ],
            'mergeFieldId': 3
        }
        r = RQS.post(
            'https://api.quickbase.com/v1/records',
            headers=headers,
            json=body
        )
        message_to_relay = 'Good'
    except Exception as e:
        message_to_relay = str(e)
    finally:
        if os.path.exists(file_name):
            os.remove(file_name)
        return message_to_relay


def insert_skunkworks_record(table_id, qb_realm, auth_token):
    message_to_relay = 'Good'
    final_auth_token = 'QB-USER-TOKEN ' + auth_token
    target_url = 'https://api.quickbase.com/v1/records'

    try:
        headers = {
            'QB-Realm-Hostname': qb_realm,
            'User-Agent': '{User-Agent}',
            'Authorization': final_auth_token
        }
        params = {
            'to': table_id,
            'data': [
              {
                '8': {
                    'value': 'test_skunkworks_table_insert'
                }
              }
            ]
        }
        r = RQS.post(
            target_url,
            headers=headers,
            json=params
        )
        # PING LOCALHOST: http://127.0.0.1:5000/ping/qb_query?table_id=bnvqs469i&fid_array=8,9&criteria={%2763%27.XEX.%272%27}
        message_to_relay = json.dumps(r.json(), indent=4)
    except Exception as e:
        message_to_relay = str(e)
    finally:
        return message_to_relay


def insert_qb_record_x(table_id, qb_realm, auth_token, field_values_array):
    message_to_relay = 'Good'
    final_auth_token = 'QB-USER-TOKEN ' + auth_token
    target_url = 'https://api.quickbase.com/v1/records'

    try:
        headers = {
            'QB-Realm-Hostname': qb_realm,
            'User-Agent': '{User-Agent}',
            'Authorization': final_auth_token
        }
        params = {
            'to': table_id,
            'data': field_values_array
        }
        r = RQS.post(
            target_url,
            headers=headers,
            json=params
        )
        # PING LOCALHOST: http://127.0.0.1:5000/ping/qb_query?table_id=bnvqs469i&fid_array=8,9&criteria={%2763%27.XEX.%272%27}
        message_to_relay = json.dumps(r.json(), indent=4)
    except Exception as e:
        message_to_relay = str(e)
    finally:
        return message_to_relay


def send_to_skunkworks_x(tab_id, rec_id, file_path, file_name, file_fid,
                         auth_token='SECRET SECRET SECRET', qb_realm='SECRET SECRET SECRET'):
    table_id = str(tab_id)
    record_id = int(rec_id)
    final_auth_token = 'QB-USER-TOKEN ' + auth_token
    message_to_relay = ''

    try:
        file_byte_content = open(file_path, 'rb').read()  # read file as byte stream
        base_64_bytes = base64.b64encode(file_byte_content)  # convert byte into "base64 byte"
        base64_string = base_64_bytes.decode('utf-8')  # convert base54 byte stream into utf-8 encoded string

        headers = {
            'QB-Realm-Hostname': qb_realm,
            'User-Agent': '{User-Agent}',
            'Authorization': final_auth_token
        }

        body = {
            'to': table_id,
            'data': [
                {
                    '3': {
                        'value': record_id
                    },
                    file_fid: {
                        'value': {
                            'fileName': file_name,
                            'data': base64_string
                        }
                    }
                }
            ],
            'mergeFieldId': 3
        }
        r = RQS.post(
            'https://api.quickbase.com/v1/records',
            headers=headers,
            json=body
        )
        message_to_relay = 'Good'
    except Exception as e:
        message_to_relay = str(e)
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)
        return message_to_relay


def transfer_to_box(qb_realm, qb_app_id, qb_table_id, qb_record_id, qb_auth_token, box_name, box_password):

    return


def get_cell_value(input_string):
    occurances = input_string.count('>')
    # 2 count is normal, no offset needed
    offset = int(occurances/2)

    correct_index = -1
    for i in range(0, offset):
        correct_index = input_string.find('>', correct_index + 1)

    left_bound = input_string.find('>', correct_index)
    right_bound = input_string.find('<', left_bound + 1)

    result = input_string[left_bound + 1:right_bound]
    return result


def strip_date_from_string(input_string):
    """Returns properly formatted date"""
    correct_date = None

    if len(input_string) == 8:
        correct_date = datetime.datetime.strptime(input_string, "%m-%d-%y")
    elif len(input_string) == 10:
        correct_date = datetime.datetime.strptime(input_string, "%m-%d-%Y")

    return correct_date


